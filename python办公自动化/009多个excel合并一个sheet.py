from openpyxl import load_workbook,Workbook
import os

def copy_data():
    wb = Workbook()
    sh = wb.active
    all_data = []
    for name in os.listdir('./销售表'): # 当前目录下的文件名称含后缀
        path = f'./销售表/{name}' # 获取完整路径
        tmp_wb = load_workbook(path)
        tmp_sh = tmp_wb.active
        for r in range(1,tmp_sh.max_row+1): # 记住时openpyxl里行列要从1开始，不是0，而右侧要max_row+1
            row_value = [] # 每行的数据汇总到一个列表里
            for c in range(1,tmp_sh.max_column+1):
                row_value.append(tmp_sh.cell(r,c).value)
            if row_value not in all_data:# 判断是否重复，没重复再加入全局数据列表
                all_data.append(row_value)
    for data in all_data:
        sh.append(data) # sheet有append()方法
    wb.save('./create_data/09数据合并.xlsx')

if __name__ == "__main__":
    copy_data()
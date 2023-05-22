# 创建excel文档
def new():
    from openpyxl import Workbook # 注意Workbook大写
    wb = Workbook() # Workbook（）类的括号别忘记了
    sh1 = wb.active
    sh2 = wb.create_sheet('数据')
    sh3 = wb.create_sheet('人员',0) # 0表示信件工作簿首位
    wb.save('./create_data/08openpyxl创建excel.xlsx')

# 设置值及样式
def set_value():
    from openpyxl.styles import Font,Alignment,colors
    bold_italic_30_font = Font(name='微软雅黑',size=30,italic=True,bold=True,color=colors.BLUE)
    bold_italic_20_font = Font(name='宋体',size=30,italic=True,bold=True,color='cd1076')
    from openpyxl import Workbook
    wb = Workbook() 
    sh1 = wb.active
    sh1['a1'] = 'hello!'
    sh1['b2'] = 'excel'
    sh1['b2'].font = bold_italic_30_font
    sh1['c3'] = 'python'
    sh1['c3'].font = bold_italic_20_font
    wb.save('./create_data/08openpyxl创建excel.xlsx')

# 批量设置值
def set_value2():
    from openpyxl import Workbook
    wb = Workbook() 
    sh1 = wb.active
    data = ['python','java','c++']
    for i,d in enumerate(data):
        sh1.cell(i+1,1).value = d
    wb.save('./create_data/08openpyxl添加数据.xlsx')


# 设置其它格式
def set_style():
    from openpyxl.styles import Alignment,colors
    from openpyxl import Workbook
    wb = Workbook() 
    sh1 = wb.active
    sh1.row_dimensions[1].height=30
    sh1.column_dimensions['a'].width=30
    data = ['python','java','c++']
    for i,d in enumerate(data):
        sh1.cell(i+1,1).value = d
        sh1.cell(i+1,1).alignment= Alignment(horizontal='right',vertical='bottom')
    wb.save('./create_data/08openpyxl添加数据.xlsx')

# 合并单元格
def set_merge():
    from openpyxl.styles import Alignment,colors
    from openpyxl import Workbook
    wb = Workbook() 
    sh1 = wb.active
    sh1.merge_cells('a1:c2') # 注意合并格子的顺序，小到大
    sh1.merge_cells('d2:e5')
    sh1['a1'] = '横向合并' # 赋值必需在合并后左上角定位格
    sh1['d2'] = '多合并'
    wb.save('./create_data/08openpyxl合并单元格.xlsx')

# 图表制作
def set_image():
    from openpyxl import Workbook
    from datetime import date
    wb = Workbook() 
    ws = wb.active # 激活第一页工作簿
    rows=[
        ['date','batch 1','batch 2','batch 3'],
        [date(2020,12,1),40,30,25],
        [date(2020,12,2),40,25,30],
        [date(2020,12,3),50,30,45],
        [date(2020,12,4),25,35,40],
        [date(2020,12,5),30,25,30],
        [date(2020,12,5),20,40,35],
    ]
    for row in rows:
        ws.append(row)
    from openpyxl.chart import LineChart,Reference
    c1 = LineChart()
    c1.title = 'line chart'
    c1.x_axis.title = 'test_number'
    c1.y_axis.title = 'size'
    # min_col,min_row从第几行几列开始读数据,max_row,max_col到几行几列未知
    data = Reference(ws,min_row=1,max_row=7,min_col=2,max_col=4) # 图标依赖数据，除去首行标题
    c1.add_data(data,titles_from_data=True)
    ws.add_chart(c1,'a9')
    wb.save('./create_data/08openpyxl设置图表.xlsx')

if __name__ == "__main__":
    # new()
    # set_value()
    # set_value2()
    # set_style()
    # set_merge()
    set_image()
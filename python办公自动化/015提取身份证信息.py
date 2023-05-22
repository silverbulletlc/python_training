from openpyxl import load_workbook
from datetime import datetime
def create_time():
    now_year = datetime.now().year # 获得当下年份
    wb = load_workbook('./base_data/person.xlsx')
    sh = wb.active
    max_column = sh.max_column # 获得现有列数量
    for i,cell in enumerate(sh['b'],start=1): # i控制次数/行数，从1开始
        pnum = cell.value
        year = pnum[6:10] # 身份证第7位开始年，字符串0开始
        month = pnum[10:12]
        day = pnum[12:14]
        # print(f'year:{year} month:{month} day:{day}')
        age = now_year - int(year)
        sh.cell(i,max_column+1).value = year
        sh.cell(i,max_column+2).value = month 
        sh.cell(i,max_column+3).value = day
        sh.cell(i,max_column+4).value = age 
    wb.save('./create_data/15提取身份证信息.xlsx')
    
if __name__ == '__main__':
    create_time()
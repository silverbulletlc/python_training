from openpyxl import load_workbook,Workbook
import os

def copy_data():
    wb = Workbook()
    for name in os.listdir('./销售表'): # 当前目录下的文件名称含后缀
        path = f'./销售表/{name}' # 获取完整路径
        tmp_wb = load_workbook(path)
        tmp_sh = tmp_wb.active
        sh = wb.create_sheet(name[:-5]) # 通过name切片（开始到倒数5），去掉后缀xlsx
        for r in range(1,tmp_sh.max_row+1): # 记住时openpyxl里行列要从1开始，不是0，而右侧要max_row+1
            row_value = [] # 每行的数据汇总到一个列表里
            for c in range(1,tmp_sh.max_column+1):
                row_value.append(tmp_sh.cell(r,c).value)
            sh.append(row_value)
    del wb['Sheet'] # 删除默认的第一个页面
    wb.save('./create_data/10数据合并.xlsx')

if __name__ == "__main__":
    copy_data()
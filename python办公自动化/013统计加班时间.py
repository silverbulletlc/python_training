from openpyxl import Workbook,load_workbook
from datetime import date

def create_excel():
    wb = Workbook()
    sh = wb.active
    rows = [ 
        ['date','姓名','打卡时间'],
        [date(2020,12,1),'梁超','19:50'],
        [date(2020,12,2),'沈旭霞','18:00'],
        [date(2020,12,3),'梁超','18:50'],
        [date(2020,12,4),'沈旭霞','17:50'],
        [date(2020,12,5),'江天若','18:10'],
        [date(2020,12,5),'周泽琳','16:50'],
    ]
    for row in rows:
        sh.append(row)
    wb.save('./base_data/打卡时间.xlsx')

def statistics():
    # 读取数据
    wb = load_workbook('./base_data/打卡时间.xlsx')
    sh = wb.active
    data = []
    for i in range(2,sh.max_row+1):
        tmp_data = []
        for j in range(1,sh.max_column+1):
            tmp_data.append(sh.cell(i,j).value)
    # 统计
        h,m = tmp_data[2].split(':') # 分拆小时和分钟
        full = int(h)*60 + int(m)
        tmp = full - 18*60 # 18点下班，计算加班时间
        tmp_data.append(tmp)
        # 处理时间问题
        tmp_data[0] = tmp_data[0].date() # 通过转换形成干净的日期，没有00：00：00
        data.append(tmp_data) # 每行末尾增加加班分钟时间
    # 保存
    n_wb = Workbook()
    n_sh = n_wb.active
    for d in data:
        n_sh.append(d)
    n_wb.save('./create_data/13统计加班时间.xlsx')

if __name__ == '__main__':
    create_excel()
    statistics()
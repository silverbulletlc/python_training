from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
def dum():
    wb = load_workbook('./base_data/打卡时间.xlsx')
    sh = wb.active
    # 存储哪一行是重复数据
    index = []
    tmp = [] # 没有重复的数据
    for i,c in enumerate(sh['b']):
        # 如果没有重复，添加
        if c.value not in tmp: # 注意是要单元格c的value，而不是直接c
            tmp.append(c.value) 
        else: # 如果重复了，记录行号
            index.append(i)
    fill = PatternFill('solid',fgColor='aeeeee')
    for i,r in enumerate(sh.rows):
        if i in index:
            for c in r:
                c.fill = fill
            print(f'第{i+1}条是重复数据')
    wb.save('./create_data/14查找重复数据.xlsx')

if __name__ == '__main__':
    dum()

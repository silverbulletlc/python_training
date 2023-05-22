from openpyxl import load_workbook,Workbook

def create_salary():
    wb = load_workbook('./base_data/工资数据.xlsx',data_only=True) # data_only表示只针对数据，可以避免因公式表达的数据缺失
    sh = wb.active
    # 按行数max_row
    title = ['序号','姓名','月工资','补贴','五险一金','所得税','罚款','奖金','最后收入']
    for i,row in enumerate(sh.rows,start=1): # 行数从1开始到末，sh.rows不是数字而是代表所有行，不用+1
        if i == 1: # 第一行为标题，调过不读取
            continue
        else:
            tmp_salary_work = Workbook()
            tmp_sh = tmp_salary_work.active
            tmp_sh.append(title) # 先加上标题行
            row_data = [cell.value for cell in row] # 简写提取每行的数据形成列表
            tmp_sh.append(row_data) 
            tmp_salary_work.save(f'./create_data/{row_data[1]}.xlsx') # row_data[1]为数据里的姓名

if __name__ == '__main__':
    create_salary()
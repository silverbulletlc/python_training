# pip3 install openpyxl

# 打开工作表方法
def open():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    # 获取工作簿
    sh1 = wb.active
    sh2 = wb['Sheet1'] # 注意“Sheet1”，s大写
    sh3 = wb.get_sheet_by_name('Sheet1')
    print(sh1 is sh2 is sh3)

# 显示工作簿名称的方法
def show_sheets():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    print(wb.sheetnames) # 打印出来的时工作簿名的序列
    for sh in wb:
        print(sh.title) # 打印出来工作簿的一个个名称

# 获取单一值
def get_one_value():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh1 = wb.active
    # 注意，openpy模块里，单元格的序号从1开始，不是从0开始
    value1 = sh1.cell(2,3).value  # （2，3）为第2行，c列
    value2 = sh1['c2'].value
    print(value1,value2)

# 获取多值
def get_many_value():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh1 = wb['Sheet1']
    # 切片
    cells1 = sh1['c2:d3']
    # print(cells1)
    # 整行，整列
    cell_row = sh1[3]
    cell_col = sh1['c']
    print(cell_row,cell_col)
    # 行切片
    cell_row3_5 = sh1[3:4]
    # 通过迭代获取数据
    for row in sh1.iter_rows(min_row=2,max_row=5,max_col=3): # 遍历行，范围时第2行到第5行，前3列
        for cell in row: # 行中遍历列数据
            print(cell.value)

# 获取全部值
def get_all_data():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh1 = wb.active

    # 先行后列
    for row in sh1.rows: # 注意这个模块里是“rows”，不是“nrows”
        for cell in row:
            print(cell.value)
    # 先列后行
    for column in sh1.columns: # 注意这个模块里时“columns”，不是“ncols”
        for cell in column:
            print(cell.value)

# 获取行列数
def get_num():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh1 = wb.active
    print(sh1.max_row)
    print(sh1.max_column)


if __name__ == "__main__":
    open()
    show_sheets()
    get_one_value()
    get_many_value()
    get_all_data()
    get_num()
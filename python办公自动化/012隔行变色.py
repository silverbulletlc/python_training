from openpyxl import Workbook
from datetime import date # 为了下文的date方法不出错需要导入
from openpyxl.styles import PatternFill # 导入填充样式
def create_excel():
    wb = Workbook()
    sh = wb.active
    rows = [ 
        ['date','batch 1','batch 2','batch 3'],
        [date(2020,12,1),40,30,25],
        [date(2020,12,2),40,25,30],
        [date(2020,12,3),50,30,45],
        [date(2020,12,4),25,35,40],
        [date(2020,12,5),30,25,30],
        [date(2020,12,5),20,40,35],
    ]
    for row in rows:
        sh.append(row)
    # 修改样式，背景色填充
    black_color = PatternFill('solid',fgColor='aeeeee')

    for i in range(1,sh.max_row+1):
        if i%2 == 0: # 偶数行填充变色
            for cell in range(1,sh.max_column+1):
                sh.cell(i,cell).fill = black_color
    wb.save('./create_data/12隔行变色.xlsx')

if __name__ == '__main__':
    create_excel()
